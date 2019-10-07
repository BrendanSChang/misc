"""TODO: comment
"""

import logging
import os
import re
from typing import List, Sequence

import lark
from openpyxl.reader.excel import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

grammar = r"""
start: import | assign | write

import: ALIAS WS* "=" WS* PATH
assign: CONTEXT WS* "=" WS* expression
write: index "=" expression

index: (CONTEXT | CNAME) ("[" expression [":" expression] "]")+
?expression: (expression WS* OP WS* expression)
          | value
          | "(" expression ")"
?value: index | CONTEXT | NUMBER | CNAME | LITERAL

ALIAS: /@\w+/
CONTEXT: /\$\w+/
LITERAL: /\"[ \w]+\"/
PATH: (LETTER | DIGIT | "-" | "_" | "/" | "\\" )+ "." ("xlsx" | "csv")
OP: "+" | "-" | "*" | "/" | "%" | "&"

%import common.WS
%import common.NUMBER
%import common.CNAME
%import common.LETTER
%import common.DIGIT
%ignore WS
"""


def load_file(path: str) -> List[str]:
  with open(path, "r") as f:
    lines = f.read().strip().split("\n")

  # Skip any comments.
  return [line.strip() for line in lines if line.strip()[0] != "#"]


class Resolver:
  """TODO: comment
  """

  TARGET_RE = re.compile(r"out(\[.+\])(\[.+\])")
  PARSER = lark.Lark(grammar, parser="lalr")

  def __init__(self, mapping: Sequence[str]):
    """TODO:comment
    """
    self._mapping = mapping
    self._output_name = None
    self._output = None
    self._sheets = {}
    self._cell_alias_tokens = {}
    self._aliases = {}
    self._local_context = {}

  def _resolve_cell_aliases(self, sheet: Worksheet, sheet_alias: str) -> None:
    cell_aliases = {}
    for row in range(sheet.min_row, sheet.max_row + 1):
      for col in range(sheet.min_column, sheet.max_column + 1):
        v = sheet.cell(row, col).value
        if v in self._cell_alias_tokens:
          cell_aliases[v] = (row, col)

    self._aliases[sheet_alias] = cell_aliases

  '''
  def _load(self, target: str, path: str) -> None:
    """TODO:comment
    """
    if target == "out":
      self._output = load_workbook(path)
      self._output_name = os.path.split(path)[1]
      self._sheets["out"] = self._output.worksheets[0]
    elif target == "alias":
      contents = load_file(path)
      for line in contents:
        values = [value.strip() for value in line.split(",")]
        if len(values) != 2:
          raise ValueError(
              "alias must have exactly two values: '{0}'".format(line))

        alias, token = values
        self._cell_alias_tokens[token] = alias

      for sheet_alias, sheet in self._sheets:
        self._resolve_cell_aliases(sheet, sheet_alias)
    else:
      raise ValueError("generic loads not yet implemented")

  def _execute(self, op: str, x: str, y: str) -> str:
    if op == "+":
      return str(int(x) + int(y))
    elif op == "-":
      return str(int(x) - int(y))
    elif op == "*":
      return str(int(x) * int(y))
    elif op == "/":
      return str(int(x) / int(y))
    elif op == "%":
      return str(int(x) % int(y))
    elif op == "&":
      return x + y
    else:
      raise ValueError("unrecognized op: {0}".format(op))

  def _nested_expression_end(
          self, expression: str, index: int, open_tag: str, close_tag: str
  ) -> int:
    count = 1
    while count > 0 and index < len(expression):
      if expression[index] == close_tag:
        count -= 1
      elif expression[index] == open_tag:
        count += 1
      index += 1

    if index == len(expression) and count > 0:
      raise ValueError(
          "could not parse nested expression '{0}'".format(expression))

    return index

  def _evaluate(self, expression):
    # Naive parsing (left-to-right).
    # +-/*%, $, (), []
    contexts = []
    op = "+"
    cur = 0
    value = 0
    i = 0
    while i < len(expression):
      if expression[i] == "(":
        start = i
        i = self._nested_expression_end(expression, i, "(", ")")
        value = self._execute(op, value, self._evaluate(expression[start+1:i]))
        cur = i + 1
      elif expression[i] == '"':
        pass
      elif expression[i] == "[":
        start = i
        i = self._nested_expression_end(expression, i, "[", "]")
        lookup = self._execute(op, cur, self._evaluate(expression[start+1:i]))
      elif expression[i] == "$":
        pass
    return value

  def _resolve_target(self, target):
    target_match = Resolver.TARGET_RE.match(target)
    if not target_match:
      raise ValueError("unable to parse target: '{0}'".format(target))

    lhs = self._evaluate(target)
  '''

  def resolve(self) -> None:
    """TODO:comment
    """
    for mapping in self._mapping:
      logging.info("resolved mapping: {0}".format(
          Resolver.PARSER.parse(mapping).pretty()))
      """
      operands = [operand.strip() for operand in mapping.split('=')]
      if len(operands) != 2:
        raise RuntimeError("invalid assignment: '{0}'".format(mapping))

      lhs, rhs = operands
      lhs = "".join([c for c in lhs if c != " "])
      rhs = "".join([c for c in rhs if c != " "])
      if lhs[0] == "@":
        self._load(lhs[1:], rhs)
      elif lhs[0] == "$":
        self._local_context[lhs[1:]] = self._evaluate(value)
      else:
        self._write(lhs, rhs)
      """

  def save(self, output_dir) -> None:
    """TODO:comment
    """
    if self._output is None:
      raise RuntimeError("unable to save because output is None")

    self._output.save(os.path.join(output_dir, self._output_name))
